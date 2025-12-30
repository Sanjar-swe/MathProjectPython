import openpyxl
from openpyxl.styles import Font, PatternFill
import random

def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Math Questions"

    # Define headers
    headers = ["Text", "Option A", "Option B", "Option C", "Option D", "Correct Answer"]
    ws.append(headers)

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    questions = []
    
    # Generate 100 questions
    for i in range(1, 101):
        num1 = random.randint(1, 100)
        num2 = random.randint(1, 100)
        operation = random.choice(['+', '-', '*', '/'])
        
        if operation == '+':
            ans = num1 + num2
            text = f"{num1} + {num2} = ?"
        elif operation == '-':
            ans = num1 - num2
            text = f"{num1} - {num2} = ?"
        elif operation == '*':
            ans = num1 * num2
            text = f"{num1} * {num2} = ?"
        else: # Division (ensure integer result)
            ans = num1
            num1 = num1 * num2
            text = f"{num1} / {num2} = ?"
            
        correct_option_char = random.choice(['a', 'b', 'c', 'd'])
        options = {}
        
        # Generate distractors
        distractors = set()
        while len(distractors) < 3:
            d = ans + random.randint(-10, 10)
            if d != ans:
                distractors.add(d)
        
        distractors = list(distractors)
        
        options[correct_option_char] = str(ans)
        remaining_chars = [c for c in ['a', 'b', 'c', 'd'] if c != correct_option_char]
        
        for idx, char in enumerate(remaining_chars):
            options[char] = str(distractors[idx])
            
        questions.append([
            text,
            options['a'],
            options['b'],
            options['c'],
            options['d'],
            correct_option_char
        ])

    for q in questions:
        ws.append(q)

    filename = "test_questions_100.xlsx"
    wb.save(filename)
    print(f"Successfully created '{filename}' with {len(questions)} questions.")

if __name__ == "__main__":
    create_excel()
